import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import pytz
from datetime import datetime
import streamlit as st
from github import Github
from bs4 import BeautifulSoup

# GitHubへのファイルアップロード関数
def upload_file_to_github(file_path, repo_name, file_name_in_repo, commit_message, GITHUB_TOKEN):
    try:
        g = Github(GITHUB_TOKEN)
        repo = g.get_repo(repo_name)
        
        with open(file_path, 'rb') as file:
            content = file.read()

        path = file_name_in_repo

        try:
            contents = repo.get_contents(path)
            repo.update_file(path, commit_message, content, contents.sha)
            st.info(f"{file_name_in_repo} を更新しました。")
        except Exception as e_inner:
            repo.create_file(path, commit_message, content)
            st.info(f"{file_name_in_repo} を作成しました。")
    except Exception as e_outer:
        st.error(f"GitHubへのファイルアップロード中にエラーが発生しました: {e_outer}")

# データ抽出と保存
def extract_data_and_save_to_csv(html_content, output_csv_path):
    soup = BeautifulSoup(html_content, "lxml")
    rows = soup.find_all("tr")[1:]

    data = {
        "台番号": [], "累計スタート": [], "BB回数": [], "RB回数": [], 
        "ART回数": [], "最大持玉": [], "BB確率": [], "RB確率": [], 
        "ART確率": [], "合成確率": []
    }

    for row in rows:
        cells = row.find_all("td")
        if len(cells) > 1:
            data["台番号"].append(cells[1].get_text())
            data["累計スタート"].append(cells[2].get_text())
            data["BB回数"].append(cells[3].get_text())
            data["RB回数"].append(cells[4].get_text())
            data["ART回数"].append(cells[5].get_text())
            data["最大持玉"].append(cells[6].get_text())
            data["BB確率"].append(cells[7].get_text())
            data["RB確率"].append(cells[8].get_text())
            data["ART確率"].append(cells[9].get_text())
            data["合成確率"].append(cells[10].get_text())

    df = pd.DataFrame(data)
    df.to_csv(output_csv_path, index=False, encoding="shift-jis")
    return df

# CSVファイルから新しいExcelファイルを作成
def create_new_excel_with_all_data(output_csv_dir, excel_path):
    csv_files = [os.path.join(output_csv_dir, f) for f in os.listdir(output_csv_dir) if f.endswith('.csv')]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合成確率"

    ws.cell(row=1, column=1, value="台番号")
    
    all_data = {}
    date_columns = []

    for csv_file in csv_files:
        df = pd.read_csv(csv_file, encoding="shift-jis")
        date = os.path.basename(csv_file).split('_')[-1].replace('.csv', '')
        formatted_date = pd.to_datetime(date).strftime('%Y/%m/%d')
        date_columns.append(formatted_date)
        
        for index, row in df.iterrows():
            if row['台番号'] not in all_data:
                all_data[row['台番号']] = {}
            all_data[row['台番号']][formatted_date] = row['合成確率']
    
    for col_index, date in enumerate(sorted(date_columns), start=2):
        ws.cell(row=1, column=col_index, value=date)

    for row_index, (machine_number, dates_data) in enumerate(all_data.items(), start=2):
        ws.cell(row=row_index, column=1, value=machine_number)
        for col_index, date in enumerate(sorted(date_columns), start=2):
            ws.cell(row=row_index, column=col_index, value=dates_data.get(date, None))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    mei_font = Font(name="メイリオ")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = mei_font

    wb.save(excel_path)

# Excelファイルに色付け
def apply_color_fill_to_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                cell_value = float(cell.value)
                if cell_value < 125:
                    cell.fill = yellow_fill
                elif 125 <= cell_value < 140:
                    cell.fill = light_blue_fill
            except (TypeError, ValueError):
                pass

    wb.save(excel_path)

# Streamlitアプリケーションのインターフェース
st.title("🎰 Juggler Data Manager 🎰")
st.write("このアプリでは、HTMLからデータを抽出し、Excelファイルに保存し、色付けします。")

# データ取得用のリンク
st.markdown("[台データオンライン メッセ武蔵境店](https://daidata.goraggio.com/100686)からデータを取得してください。")

# GitHubトークンの取得
GITHUB_TOKEN = st.secrets["github"]["token"]

# 日本時間の今日の日付を取得
japan_time_zone = pytz.timezone('Asia/Tokyo')
current_date_japan = datetime.now(japan_time_zone)

# HTMLファイルのアップロードまたは貼り付け
uploaded_html = st.file_uploader("HTMLファイルをアップロード", type=["html", "htm", "txt"])
html_input = st.text_area("または、直接HTMLを貼り付け", "")
date_input = st.date_input("日付を選択", current_date_japan)

# ファイルの処理開始ボタン
if st.button("処理開始"):
    if uploaded_html or html_input:
        if uploaded_html:
            html_content = uploaded_html.getvalue().decode("utf-8")
        else:
            html_content = html_input
        
        output_csv_dir = "./マイジャグラーV"
        if not os.path.exists(output_csv_dir):
            os.makedirs(output_csv_dir)

        output_csv_path = os.path.join(output_csv_dir, f"slot_machine_data_{date_input}.csv")
        excel_path = "マイジャグラーV_塗りつぶし済み.xlsx"
        
        # データ処理とExcelファイル作成
        df_new = extract_data_and_save_to_csv(html_content, output_csv_path)
        create_new_excel_with_all_data(output_csv_dir, excel_path)
        apply_color_fill_to_excel(excel_path)

        st.success(f"データ処理が完了し、{excel_path} に保存されました。")

        # GitHubアップロード
        repo_name = "yudai4452/juggler-data-apps"
        commit_message = f"Add data for {date_input}"
        upload_file_to_github(excel_path, repo_name, excel_path, commit_message, GITHUB_TOKEN)
        upload_file_to_github(output_csv_path, repo_name, f"マイジャグラーV/{os.path.basename(output_csv_path)}", commit_message, GITHUB_TOKEN)

        # ダウンロードリンクの表示
        st.download_button(label="CSVファイルをダウンロード", data=open(output_csv_path, 'rb').read(), file_name=f"slot_machine_data_{date_input}.csv")
        st.download_button(label="Excelファイルをダウンロード", data=open(excel_path, 'rb').read(), file_name="マイジャグラーV_塗りつぶし済み.xlsx")

        # 可視化アプリへのリンク
        st.markdown("[こちらをクリックしてJuggler Data Visualizerへ移動](https://juggler-data-apps-6qz2wrn69bezyvzykh5bdb.streamlit.app/)")

# 可視化アプリへのリンク
st.markdown("[こちらをクリックしてJuggler Data Visualizerへ移動](https://juggler-data-apps-6qz2wrn69bezyvzykh5bdb.streamlit.app/)")
